#!/usr/bin/env python3
"""PINNACLE LGS — PDF & Excel generator (V12)"""

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import (
    BaseDocTemplate, Frame, PageTemplate, Table, TableStyle,
    Paragraph, Spacer, KeepTogether, Image, PageBreak, NextPageTemplate
)
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_RIGHT, TA_CENTER
from PIL import Image as PILImage
import os

ORANGE       = colors.HexColor("#E8841A")
ORANGE_LIGHT = colors.HexColor("#FDF0E2")
ORANGE_MED   = colors.HexColor("#F5A94A")
DARK         = colors.HexColor("#1A1A1A")
DARK2        = colors.HexColor("#2C2C2C")
GRAY         = colors.HexColor("#555555")
GRAY_L       = colors.HexColor("#888888")
SILVER       = colors.HexColor("#F0F0F0")
WHITE        = colors.white
GRID         = colors.HexColor("#E0E0E0")
ROW_ALT      = colors.HexColor("#FFF8F2")

# Assets sont dans le dossier assets/ à côté du code
ASSETS = os.path.join(os.path.dirname(os.path.abspath(__file__)), "assets")
LOGO          = os.path.join(ASSETS, "Pinnacle_logo.png")
SIG           = os.path.join(ASSETS, "Signature_Antoine.jpg")
COVER_PHOTO   = os.path.join(ASSETS, "photo_building.jpeg")
MACHINE_PHOTO = os.path.join(ASSETS, "photo_machine.png")

W, H = A4
L_MARGIN = 15*mm
R_MARGIN = 12*mm
CW = W - L_MARGIN - R_MARGIN

_GRADIENT_PATH = None

def fmt(n):
    return f"${n:,.0f}" if not isinstance(n, str) else n

def get_gradient_png():
    global _GRADIENT_PATH
    if _GRADIENT_PATH is None or not os.path.exists(_GRADIENT_PATH):
        path = "/tmp/_pinnacle_gradient.png"
        w_px, h_px = 200, 1200
        img = PILImage.new('RGBA', (w_px, h_px), (0, 0, 0, 0))
        pixels = img.load()
        for y in range(h_px):
            progress = y / h_px
            alpha = int((0.04 + progress * 0.36) * 255)
            for x in range(w_px):
                pixels[x, y] = (0, 0, 0, alpha)
        img.save(path)
        _GRADIENT_PATH = path
    return _GRADIENT_PATH



# ============================================================
# LEAD TIMES (source : Template_proforma_202410_En.xlsx)
# Lead time = max parmi toutes les machines demandées
# ============================================================
LEADTIME_MAP = {
    "X1":     "3 to 4 months",
    "X1_18":  "3 to 4 months",
    "X1_16":  "3 to 4 months",
    "X1_20":  "3 to 4 months",
    "X2":     "3 to 4 months",
    "X2_12":  "3 to 4 months",
    "X2_16":  "3 to 4 months",
    "X20":    "3 to 4 months",
    "X30I":   "3 to 4 months",
    "X3":     "3 to 4 months",
    "X3I":    "3 to 4 months",
    "X5":     "4 to 5 months",
    "X8I":    "4 to 5 months",
    "X80I":   "4 to 5 months",
    "X6":     "6 to 7 months",
    "X10I":   "6 to 7 months",
    "X100I":  "6 to 7 months",
    "X88":    "6 to 7 months",
    "X88I":   "6 to 7 months",
    "X888":   "8 to 10 months",
    "X888I":  "8 to 10 months",
    "X168":   "10 to 12 months",
}

LEADTIME_RANK = {
    "3 to 4 months":  1,
    "4 to 5 months":  2,
    "6 to 7 months":  3,
    "8 to 10 months": 4,
    "10 to 12 months": 5,
}

def compute_lead_time(machines):
    """Returns the maximum lead time among all machines in the quote"""
    if not machines:
        return "3 to 4 months"
    best_lt = "3 to 4 months"
    best_rank = 1
    for m in machines:
        code = (m.get("model") or "").upper()
        lt = LEADTIME_MAP.get(code)
        if lt:
            r = LEADTIME_RANK.get(lt, 1)
            if r > best_rank:
                best_rank = r
                best_lt = lt
    return best_lt


class QuoteDoc(BaseDocTemplate):
    def __init__(self, fn, quote_fn="", **kw):
        self.quote_fn = quote_fn
        super().__init__(fn, **kw)


def cover_cb(canvas_obj, doc):
    d = doc._cover_data
    if os.path.exists(COVER_PHOTO):
        canvas_obj.drawImage(COVER_PHOTO, 0, 0, W, H,
                             preserveAspectRatio=False, mask='auto')
    canvas_obj.drawImage(get_gradient_png(), 0, 0, W, H, mask='auto')

    if os.path.exists(LOGO):
        canvas_obj.drawImage(LOGO, L_MARGIN, H-36*mm,
                             width=62*mm, height=17*mm,
                             preserveAspectRatio=True, mask='auto')

    canvas_obj.setFillColor(WHITE)
    canvas_obj.setFont("Helvetica-Bold", 9)
    canvas_obj.drawRightString(W-R_MARGIN, H-18*mm, f"Ref: {d['reference']}")
    canvas_obj.setFont("Helvetica", 8.5)
    canvas_obj.drawRightString(W-R_MARGIN, H-25*mm, d['date'])

    canvas_obj.setFillColor(WHITE)
    canvas_obj.setFont("Helvetica-Bold", 52)
    canvas_obj.drawString(L_MARGIN, H*0.46, d.get('document_type', 'PROFORMA'))

    canvas_obj.setFillColor(ORANGE)
    canvas_obj.roundRect(L_MARGIN, H*0.405, 94*mm, 8*mm, 2*mm, fill=True, stroke=False)
    canvas_obj.setFillColor(WHITE)
    canvas_obj.setFont("Helvetica-Bold", 8.5)
    canvas_obj.drawString(L_MARGIN+4*mm, H*0.405+2.5*mm, "LIGHT GAUGE STEEL FRAMING SYSTEMS")

    models = "  ·  ".join([m["model"] for m in d["machines"]])
    canvas_obj.setFillColor(colors.Color(1,1,1,alpha=0.15))
    canvas_obj.roundRect(L_MARGIN, H*0.35, W-L_MARGIN-R_MARGIN, 7*mm, 1*mm, fill=True, stroke=False)
    canvas_obj.setFillColor(ORANGE_MED)
    canvas_obj.setFont("Helvetica-Bold", 8.5)
    canvas_obj.drawString(L_MARGIN+4*mm, H*0.35+2*mm, f"Systems:  {models}")

    cl = d['client']
    canvas_obj.setFillColor(colors.Color(0,0,0,alpha=0.42))
    canvas_obj.roundRect(L_MARGIN, H*0.13, W-L_MARGIN-R_MARGIN, 52*mm, 3*mm, fill=True, stroke=False)
    canvas_obj.setFillColor(ORANGE)
    canvas_obj.setFont("Helvetica-Bold", 7.5)
    canvas_obj.drawString(L_MARGIN+6*mm, H*0.13+42*mm, "PREPARED FOR")
    canvas_obj.setFillColor(WHITE)
    canvas_obj.setFont("Helvetica-Bold", 20)
    canvas_obj.drawString(L_MARGIN+6*mm, H*0.13+30*mm, cl['company'])
    canvas_obj.setFont("Helvetica", 11)
    canvas_obj.drawString(L_MARGIN+6*mm, H*0.13+21*mm, cl['name'])
    canvas_obj.setFillColor(colors.HexColor("#DDDDDD"))
    canvas_obj.setFont("Helvetica", 9)
    canvas_obj.drawString(L_MARGIN+6*mm, H*0.13+14*mm, cl['country'])
    canvas_obj.drawString(L_MARGIN+6*mm, H*0.13+8*mm,  cl['email'])
    canvas_obj.drawString(L_MARGIN+6*mm, H*0.13+2*mm,  cl['phone'])

    canvas_obj.setFillColor(ORANGE)
    canvas_obj.rect(0, 0, W, 11*mm, fill=True, stroke=False)
    canvas_obj.setFillColor(WHITE)
    canvas_obj.setFont("Helvetica-Bold", 7.5)
    canvas_obj.drawString(L_MARGIN, 3.8*mm,
        "Pinnacle LGS Inc.  ·  1636 240th Street, Harbor City, CA 90710, USA")
    canvas_obj.drawRightString(W-R_MARGIN, 3.8*mm, "www.pinnaclelgs.com")


def inner_cb(canvas_obj, doc):
    canvas_obj.saveState()
    if os.path.exists(LOGO):
        canvas_obj.drawImage(LOGO, L_MARGIN, H-22*mm,
                             width=44*mm, height=12*mm,
                             preserveAspectRatio=True, mask='auto')
    canvas_obj.setFillColor(GRAY)
    canvas_obj.setFont("Helvetica", 6.5)
    for i, line in enumerate([
        "Pinnacle LGS Inc.",
        "1636 240th Street, Harbor City, CA 90710, USA",
        "www.pinnaclelgs.com"
    ]):
        canvas_obj.drawRightString(W-R_MARGIN, H-9.5*mm - i*3.3*mm, line)
    canvas_obj.setStrokeColor(colors.HexColor("#CCCCCC"))
    canvas_obj.setLineWidth(0.5)
    canvas_obj.line(L_MARGIN, H-25*mm, W-R_MARGIN, H-25*mm)
    canvas_obj.setLineWidth(0.4)
    canvas_obj.line(L_MARGIN, 13*mm, W-R_MARGIN, 13*mm)
    canvas_obj.setFillColor(GRAY_L)
    canvas_obj.setFont("Helvetica", 5.8)
    canvas_obj.drawString(L_MARGIN, 8*mm, doc.quote_fn)
    canvas_obj.setFillColor(ORANGE)
    canvas_obj.setFont("Helvetica-Bold", 7)
    canvas_obj.drawRightString(W-R_MARGIN, 8*mm, f"Page {canvas_obj.getPageNumber()}")
    canvas_obj.restoreState()


def build_pdf(data, out, fn=""):
    # Override lead_time with computed value based on machines
    data = dict(data)
    data["lead_time"] = compute_lead_time(data.get("machines", []))
    doc = QuoteDoc(out, quote_fn=fn or os.path.basename(out),
                   pagesize=A4,
                   topMargin=30*mm, bottomMargin=20*mm,
                   leftMargin=L_MARGIN, rightMargin=R_MARGIN)
    doc._cover_data = data

    doc.addPageTemplates([
        PageTemplate(id='cover',
                     frames=[Frame(0,0,W,H,id='cv',leftPadding=0,rightPadding=0,
                                   topPadding=0,bottomPadding=0)],
                     onPage=cover_cb),
        PageTemplate(id='inner',
                     frames=[Frame(L_MARGIN, 20*mm, CW, H-30*mm-20*mm, id='main',
                                   leftPadding=0, rightPadding=0,
                                   topPadding=0, bottomPadding=0)],
                     onPage=inner_cb),
    ])

    sn  = ParagraphStyle('N', fontSize=8.5, textColor=DARK2, fontName='Helvetica', leading=12)
    ss  = ParagraphStyle('S', fontSize=7.5, textColor=GRAY,  fontName='Helvetica', leading=10)
    sb  = ParagraphStyle('B', fontSize=8.5, textColor=DARK2, fontName='Helvetica-Bold', leading=12)
    sr  = ParagraphStyle('R', fontSize=8.5, textColor=DARK2, fontName='Helvetica', alignment=TA_RIGHT)
    srb = ParagraphStyle('RB',fontSize=8.5, textColor=DARK2, fontName='Helvetica-Bold', alignment=TA_RIGHT)
    sc  = ParagraphStyle('C', fontSize=8.5, textColor=DARK2, fontName='Helvetica', alignment=TA_CENTER)

    def sh(text):
        t = Table([[Paragraph(f"<b>{text}</b>",
                   ParagraphStyle('SH',fontSize=9.5,textColor=WHITE,fontName='Helvetica-Bold'))]],
                  colWidths=[CW])
        t.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(-1,-1),ORANGE),
            ('TOPPADDING',(0,0),(-1,-1),5),
            ('BOTTOMPADDING',(0,0),(-1,-1),5),
            ('LEFTPADDING',(0,0),(-1,-1),8),
        ]))
        return t

    C1 = CW - 18*mm - 38*mm - 38*mm
    COL = [C1, 18*mm, 38*mm, 38*mm]

    def th_row():
        s   = ParagraphStyle('TH', fontSize=8.5,textColor=WHITE,fontName='Helvetica-Bold')
        sr2 = ParagraphStyle('THR',fontSize=8.5,textColor=WHITE,fontName='Helvetica-Bold',alignment=TA_RIGHT)
        sc2 = ParagraphStyle('THC',fontSize=8.5,textColor=WHITE,fontName='Helvetica-Bold',alignment=TA_CENTER)
        return [Paragraph("Description",s), Paragraph("Qty",sc2),
                Paragraph("Unit Price (USD)",sr2), Paragraph("Total (USD)",sr2)]

    def dt(rows):
        t = Table(rows, colWidths=COL)
        ts = [
            ('BACKGROUND',(0,0),(-1,0),ORANGE),('TEXTCOLOR',(0,0),(-1,0),WHITE),
            ('BOX',(0,0),(-1,-1),0.5,GRID),('INNERGRID',(0,0),(-1,-1),0.3,GRID),
            ('TOPPADDING',(0,0),(-1,-1),5),('BOTTOMPADDING',(0,0),(-1,-1),5),
            ('LEFTPADDING',(0,0),(-1,-1),6),('RIGHTPADDING',(0,0),(-1,-1),6),
            ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
        ]
        for i in range(1,len(rows)):
            if i%2==0: ts.append(('BACKGROUND',(0,i),(-1,i),ROW_ALT))
        t.setStyle(TableStyle(ts))
        return t

    story = []
    story.append(Spacer(1,1))
    story.append(NextPageTemplate('inner'))
    story.append(PageBreak())

    story.append(Paragraph(f"<b>{data.get('document_type', 'PROFORMA')}</b>",
        ParagraphStyle('TIT',fontSize=22,textColor=DARK,fontName='Helvetica-Bold',
                       spaceAfter=5*mm)))

    W1,W2,W3,W4 = 28*mm, CW/2-28*mm, 28*mm, CW/2-28*mm
    meta = [
        [Paragraph("<b>Reference:</b>",sb), Paragraph(data["reference"],sn),
         Paragraph("<b>Date:</b>",sb), Paragraph(data["date"],sn)],
        [Paragraph("<b>Sales Person:</b>",sb), Paragraph(data["sales_person"],sn),
         Paragraph("<b>Validity:</b>",sb), Paragraph(data["validity"],sn)],
        [Paragraph("<b>Delivery Terms:</b>",sb), Paragraph(data["delivery_terms"],sn),
         Paragraph("<b>Lead Time:</b>",sb), Paragraph(data["lead_time"],sn)],
    ]
    mt = Table(meta, colWidths=[W1,W2,W3,W4])
    mt.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,-1),SILVER),
        ('BOX',(0,0),(-1,-1),0.5,GRID),('INNERGRID',(0,0),(-1,-1),0.3,GRID),
        ('TOPPADDING',(0,0),(-1,-1),4),('BOTTOMPADDING',(0,0),(-1,-1),4),
        ('LEFTPADDING',(0,0),(-1,-1),5),
    ]))
    story.append(mt)
    story.append(Spacer(1,4*mm))

    cl = data["client"]
    ch = ParagraphStyle('CH',fontSize=8.5,textColor=WHITE,fontName='Helvetica-Bold')
    PHOTO_W = 72*mm
    GAP     = 3*mm
    CT_W    = CW - PHOTO_W - GAP
    CT_C1, CT_C2 = 24*mm, CT_W - 24*mm

    ct = Table([
        [Paragraph("BILL TO",ch),""],
        [Paragraph("<b>Company:</b>",sb), Paragraph(cl['company'],sn)],
        [Paragraph("<b>Country:</b>",sb), Paragraph(cl['country'],sn)],
        [Paragraph("<b>Contact:</b>",sb), Paragraph(cl['name'],sn)],
        [Paragraph("<b>Email:</b>",sb),   Paragraph(cl['email'],sn)],
        [Paragraph("<b>Phone:</b>",sb),   Paragraph(cl['phone'],sn)],
    ], colWidths=[CT_C1, CT_C2])
    ct.setStyle(TableStyle([
        ('SPAN',(0,0),(-1,0)),('BACKGROUND',(0,0),(-1,0),ORANGE),
        ('BOX',(0,0),(-1,-1),0.5,GRID),('INNERGRID',(0,1),(-1,-1),0.3,GRID),
        ('TOPPADDING',(0,0),(-1,-1),4),('BOTTOMPADDING',(0,0),(-1,-1),4),
        ('LEFTPADDING',(0,0),(-1,-1),6),
    ]))

    if os.path.exists(MACHINE_PHOTO):
        m_img = Image(MACHINE_PHOTO, width=PHOTO_W, height=35*mm)
    else:
        m_img = Spacer(PHOTO_W, 35*mm)

    m_lbl = Paragraph(
        "<font color='#E8841A'><b>World's Most Advanced</b></font><br/>"
        "<b>Cold Formed Steel Building System</b>",
        ParagraphStyle('ML',fontSize=7,textColor=DARK2,fontName='Helvetica',
                       alignment=TA_CENTER,leading=10))
    photo_col = Table([[m_img],[m_lbl]], colWidths=[PHOTO_W])
    photo_col.setStyle(TableStyle([
        ('TOPPADDING',(0,0),(-1,-1),0),('BOTTOMPADDING',(0,0),(-1,-1),2),
        ('ALIGN',(0,0),(-1,-1),'CENTER'),
    ]))

    combo = Table([[ct, "", photo_col]], colWidths=[CT_W, GAP, PHOTO_W])
    combo.setStyle(TableStyle([
        ('VALIGN',(0,0),(-1,-1),'TOP'),
        ('LEFTPADDING',(0,0),(-1,-1),0),('RIGHTPADDING',(0,0),(-1,-1),0),
        ('TOPPADDING',(0,0),(-1,-1),0),('BOTTOMPADDING',(0,0),(-1,-1),0),
    ]))
    story.append(combo)
    story.append(Spacer(1,6*mm))

    mr, mt2 = [th_row()], 0
    for m in data["machines"]:
        t = m["qty"]*m["price"]; mt2+=t
        mr.append([
            Paragraph(f"<b>{m['desc']}</b><br/>"
                      f"<font size='7' color='#888888'>Material Thickness {m['spec']}</font>",sn),
            Paragraph(str(m["qty"]),sc), Paragraph(fmt(m["price"]),sr), Paragraph(fmt(t),srb)])
    story.append(KeepTogether([sh("MACHINES"), dt(mr), Spacer(1,4*mm)]))

    sr2_, st2 = [th_row()], 0
    for s in data["software"]:
        t=s["qty"]*s["price"]; st2+=t
        sr2_.append([
            Paragraph(f"<b>{s['desc']}</b><br/>"
                      f"<font size='7' color='#888888'>{s['spec']}</font>",sn),
            Paragraph(str(s["qty"]),sc), Paragraph(fmt(s["price"]),sr), Paragraph(fmt(t),srb)])
    story.append(KeepTogether([sh("SOFTWARE & TRAINING"), dt(sr2_), Spacer(1,4*mm)]))

    cr2, ct2 = [th_row()], 0
    for ci in data["commissioning"]:
        t=ci["qty"]*ci["price"]; ct2+=t
        wk=f"{ci['qty']} week{'s' if ci['qty']>1 else ''} on-site (40h/week)"
        cr2.append([
            Paragraph(f"<b>{ci['desc']}</b><br/>"
                      f"<font size='7' color='#888888'>{wk}</font>",sn),
            Paragraph(str(ci["qty"]),sc), Paragraph(fmt(ci["price"]),sr), Paragraph(fmt(t),srb)])
    story.append(KeepTogether([sh("COMMISSIONING & INSTALLATION"), dt(cr2), Spacer(1,4*mm)]))

    inc=[sh("INCLUDED AT NO ADDITIONAL CHARGE")]
    for item in data["included"]:
        inc.append(Paragraph(
            f"<font color='#E8841A'><b>&#10003;</b></font>&nbsp;&nbsp;{item}",
            ParagraphStyle('IL',fontSize=8.5,textColor=DARK2,fontName='Helvetica',
                           leftIndent=6*mm,spaceBefore=2*mm)))
    inc.append(Spacer(1,5*mm))
    story.append(KeepTogether(inc))

    disc=0
    TW1, TW2 = CW-50*mm, 50*mm
    trows=[
        [Paragraph("<b>Subtotal — Machines</b>",sb),            Paragraph(fmt(mt2),srb)],
        [Paragraph("<b>Subtotal — Software & Training</b>",sb), Paragraph(fmt(st2),srb)],
        [Paragraph("<b>Subtotal — Commissioning</b>",sb),       Paragraph(fmt(ct2),srb)],
    ]
    if data.get("discount_pct",0)>0:
        disc=int(mt2*data["discount_pct"]/100)
        red =ParagraphStyle('D', fontSize=8.5,textColor=colors.HexColor("#CC0000"),fontName='Helvetica-Bold')
        redr=ParagraphStyle('DR',fontSize=8.5,textColor=colors.HexColor("#CC0000"),
                             fontName='Helvetica-Bold',alignment=TA_RIGHT)
        trows.append([
            Paragraph(f"<b>{data['discount_label']} ({data['discount_pct']}% on machines only)</b>",red),
            Paragraph(f"-{fmt(disc)}",redr)])
    grand=mt2-disc+st2+ct2
    trows.append([
        Paragraph("<b>TOTAL DUE (USD)</b>",
            ParagraphStyle('GT',fontSize=13,textColor=DARK,fontName='Helvetica-Bold')),
        Paragraph(f"<b>{fmt(grand)}</b>",
            ParagraphStyle('GTR',fontSize=13,textColor=ORANGE,
                           fontName='Helvetica-Bold',alignment=TA_RIGHT))])
    tt=Table(trows,colWidths=[TW1,TW2])
    tt.setStyle(TableStyle([
        ('BOX',(0,0),(-1,-1),0.5,GRID),('INNERGRID',(0,0),(-1,-1),0.3,GRID),
        ('BACKGROUND',(0,-1),(-1,-1),ORANGE_LIGHT),
        ('TOPPADDING',(0,0),(-1,-1),6),('BOTTOMPADDING',(0,0),(-1,-1),6),
        ('LEFTPADDING',(0,0),(-1,-1),8),('RIGHTPADDING',(0,0),(-1,-1),8),
    ]))
    story.append(KeepTogether([sh("SUMMARY"),tt,Spacer(1,5*mm)]))

    ph =ParagraphStyle('PH', fontSize=8.5,textColor=WHITE,fontName='Helvetica-Bold')
    phr=ParagraphStyle('PHR',fontSize=8.5,textColor=WHITE,fontName='Helvetica-Bold',alignment=TA_RIGHT)
    phc=ParagraphStyle('PHC',fontSize=8.5,textColor=WHITE,fontName='Helvetica-Bold',alignment=TA_CENTER)
    PW1,PW2,PW3,PW4 = 55*mm, 22*mm, 55*mm, CW-132*mm
    pr=[[Paragraph("Installment",ph),Paragraph("%",phc),
         Paragraph("Amount (USD)",phr),Paragraph("Due",phc)],
        [Paragraph("First deposit",sn),Paragraph("50%",sc),
         Paragraph(fmt(grand*0.5),sr),Paragraph("At the order",sc)],
        [Paragraph("Final payment",sn),Paragraph("50%",sc),
         Paragraph(fmt(grand*0.5),sr),Paragraph("Before shipment",sc)]]
    pt=Table(pr,colWidths=[PW1,PW2,PW3,PW4])
    pt.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0),ORANGE),('TEXTCOLOR',(0,0),(-1,0),WHITE),
        ('BOX',(0,0),(-1,-1),0.5,GRID),('INNERGRID',(0,0),(-1,-1),0.3,GRID),
        ('TOPPADDING',(0,0),(-1,-1),5),('BOTTOMPADDING',(0,0),(-1,-1),5),
        ('LEFTPADDING',(0,0),(-1,-1),6),('VALIGN',(0,0),(-1,-1),'MIDDLE'),
    ]))
    story.append(KeepTogether([sh("PAYMENT SCHEDULE"),pt,Spacer(1,5*mm)]))

    blk=[sh("NOTES & CONDITIONS")]
    for n in [
        "Prices are FOB Taiwan.",
        "Air tickets, accommodation, meals and local transportation for commissioning must be covered by the customer.",
        f"This quotation is valid for {data['validity']} from the date of issue.",
        f"Lead time: {data['lead_time']} after first deposit (excluding shipment).",
        "Hitachi Printer: CBS (Cross Border Sales) agreement with local Hitachi dealer is mandatory.",
        "Please refer to the attached Sales Conditions document for full terms and conditions.",
    ]:
        blk.append(Paragraph(
            f"<font color='#E8841A'><b>&#8226;</b></font>&nbsp;&nbsp;{n}",
            ParagraphStyle('NI',fontSize=8,textColor=DARK2,fontName='Helvetica',
                           leftIndent=6*mm,spaceBefore=2*mm,leading=11)))
    blk.append(Spacer(1,8*mm))
    blk.append(sh("SIGNATURES"))
    blk.append(Spacer(1,4*mm))

    sig_img = Image(SIG,width=40*mm,height=17*mm) if os.path.exists(SIG) else Spacer(40*mm,17*mm)
    SW = CW/2
    sig_t=Table([
        [Paragraph("<b>For and on behalf of<br/>Pinnacle Building Technology FZE</b>",sn),
         Paragraph(f"<b>Accepted by<br/>{cl['company']}</b>",sn)],
        [sig_img, Spacer(1,17*mm)],
        [Paragraph("________________________________<br/><b>Antoine Coelho</b><br/>"
                   "Sales Director Europe & North Africa",ss),
         Paragraph(f"________________________________<br/><b>{cl['name']}</b><br/>{cl['company']}",ss)],
    ],colWidths=[SW,SW])
    sig_t.setStyle(TableStyle([
        ('TOPPADDING',(0,0),(-1,-1),3),('BOTTOMPADDING',(0,0),(-1,-1),3),
        ('VALIGN',(0,0),(-1,-1),'BOTTOM'),
    ]))
    blk.append(sig_t)
    story.append(KeepTogether(blk))

    doc.build(story)


def build_excel(data, out):
    from openpyxl import Workbook
    # Override lead_time with computed value based on machines
    data = dict(data)
    data["lead_time"] = compute_lead_time(data.get("machines", []))
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    orange_fill = PatternFill(start_color="E8841A", end_color="E8841A", fill_type="solid")
    light_orange_fill = PatternFill(start_color="FDF0E2", end_color="FDF0E2", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True, size=11)
    bold_font = Font(bold=True, size=10)
    normal_font = Font(size=10)
    big_bold = Font(bold=True, size=14, color="E8841A")
    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    ws = wb.active
    ws.title = "Quote"
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 32
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18

    row = 1
    ws.cell(row=row, column=2, value=f"PINNACLE LGS — {data.get('document_type', 'PROFORMA')}").font = Font(bold=True, size=18, color="E8841A")
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    row += 2

    ws.cell(row=row, column=2, value="Reference").font = bold_font
    ws.cell(row=row, column=3, value=data['reference']).font = normal_font
    ws.cell(row=row, column=5, value="Date").font = bold_font
    ws.cell(row=row, column=6, value=data['date']).font = normal_font
    row += 1
    ws.cell(row=row, column=2, value="Sales Person").font = bold_font
    ws.cell(row=row, column=3, value=data['sales_person']).font = normal_font
    ws.cell(row=row, column=5, value="Validity").font = bold_font
    ws.cell(row=row, column=6, value=data['validity']).font = normal_font
    row += 1
    ws.cell(row=row, column=2, value="Delivery Terms").font = bold_font
    ws.cell(row=row, column=3, value=data['delivery_terms']).font = normal_font
    ws.cell(row=row, column=5, value="Lead Time").font = bold_font
    ws.cell(row=row, column=6, value=data['lead_time']).font = normal_font
    row += 2

    ws.cell(row=row, column=2, value="BILL TO").font = white_font
    ws.cell(row=row, column=2).fill = orange_fill
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    row += 1
    cl = data['client']
    for label, value in [("Company", cl['company']), ("Country", cl['country']),
                         ("Contact", cl['name']), ("Email", cl['email']), ("Phone", cl['phone'])]:
        ws.cell(row=row, column=2, value=label).font = bold_font
        ws.cell(row=row, column=3, value=value).font = normal_font
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
        row += 1
    row += 1

    ws.cell(row=row, column=2, value="MACHINES").font = white_font
    ws.cell(row=row, column=2).fill = orange_fill
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    row += 1
    for col, h in enumerate(["Description", "Specification", "Qty", "Unit Price (USD)", "Total (USD)"], start=2):
        c = ws.cell(row=row, column=col, value=h)
        c.font = white_font; c.fill = orange_fill; c.alignment = center; c.border = border
    row += 1
    m_total = 0
    for m in data['machines']:
        total = m['qty']*m['price']; m_total += total
        ws.cell(row=row, column=2, value=m['desc']).border = border
        ws.cell(row=row, column=3, value=f"Material Thickness {m['spec']}").border = border
        ws.cell(row=row, column=4, value=m['qty']).alignment = center; ws.cell(row=row, column=4).border = border
        ws.cell(row=row, column=5, value=m['price']).number_format = '"$"#,##0'; ws.cell(row=row, column=5).border = border
        ws.cell(row=row, column=6, value=total).number_format = '"$"#,##0'; ws.cell(row=row, column=6).font = bold_font; ws.cell(row=row, column=6).border = border
        row += 1
    row += 1

    ws.cell(row=row, column=2, value="SOFTWARE & TRAINING").font = white_font
    ws.cell(row=row, column=2).fill = orange_fill
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    row += 1
    for col, h in enumerate(["Description", "Specification", "Qty", "Unit Price (USD)", "Total (USD)"], start=2):
        c = ws.cell(row=row, column=col, value=h)
        c.font = white_font; c.fill = orange_fill; c.alignment = center; c.border = border
    row += 1
    s_total = 0
    for s in data['software']:
        total = s['qty']*s['price']; s_total += total
        ws.cell(row=row, column=2, value=s['desc']).border = border
        ws.cell(row=row, column=3, value=s['spec']).border = border
        ws.cell(row=row, column=4, value=s['qty']).alignment = center; ws.cell(row=row, column=4).border = border
        ws.cell(row=row, column=5, value=s['price']).number_format = '"$"#,##0'; ws.cell(row=row, column=5).border = border
        ws.cell(row=row, column=6, value=total).number_format = '"$"#,##0'; ws.cell(row=row, column=6).font = bold_font; ws.cell(row=row, column=6).border = border
        row += 1
    row += 1

    ws.cell(row=row, column=2, value="COMMISSIONING & INSTALLATION").font = white_font
    ws.cell(row=row, column=2).fill = orange_fill
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    row += 1
    for col, h in enumerate(["Description", "Weeks", "Qty", "Price/Week", "Total (USD)"], start=2):
        c = ws.cell(row=row, column=col, value=h)
        c.font = white_font; c.fill = orange_fill; c.alignment = center; c.border = border
    row += 1
    c_total = 0
    for ci in data['commissioning']:
        total = ci['qty']*ci['price']; c_total += total
        ws.cell(row=row, column=2, value=ci['desc']).border = border
        ws.cell(row=row, column=3, value=f"{ci['qty']} week(s)").border = border
        ws.cell(row=row, column=4, value=ci['qty']).alignment = center; ws.cell(row=row, column=4).border = border
        ws.cell(row=row, column=5, value=ci['price']).number_format = '"$"#,##0'; ws.cell(row=row, column=5).border = border
        ws.cell(row=row, column=6, value=total).number_format = '"$"#,##0'; ws.cell(row=row, column=6).font = bold_font; ws.cell(row=row, column=6).border = border
        row += 1
    row += 1

    ws.cell(row=row, column=2, value="SUMMARY").font = white_font
    ws.cell(row=row, column=2).fill = orange_fill
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    row += 1
    disc = 0
    if data.get('discount_pct', 0) > 0:
        disc = int(m_total * data['discount_pct'] / 100)
    grand = m_total - disc + s_total + c_total

    for lbl, val in [
        ("Subtotal — Machines", m_total),
        ("Subtotal — Software & Training", s_total),
        ("Subtotal — Commissioning", c_total),
    ]:
        ws.cell(row=row, column=2, value=lbl).font = bold_font
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        ws.cell(row=row, column=6, value=val).number_format = '"$"#,##0'
        ws.cell(row=row, column=6).font = bold_font
        row += 1
    if disc > 0:
        ws.cell(row=row, column=2, value=f"{data['discount_label']} ({data['discount_pct']}% on machines only)").font = Font(bold=True, color="CC0000")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        ws.cell(row=row, column=6, value=-disc).number_format = '"$"#,##0'
        ws.cell(row=row, column=6).font = Font(bold=True, color="CC0000")
        row += 1
    ws.cell(row=row, column=2, value="TOTAL DUE (USD)").font = big_bold
    ws.cell(row=row, column=2).fill = light_orange_fill
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    ws.cell(row=row, column=6, value=grand).number_format = '"$"#,##0'
    ws.cell(row=row, column=6).font = big_bold
    ws.cell(row=row, column=6).fill = light_orange_fill

    # PAYMENTS sheet
    ws2 = wb.create_sheet("Payments")
    ws2.column_dimensions['A'].width = 3
    ws2.column_dimensions['B'].width = 20
    ws2.column_dimensions['C'].width = 15
    ws2.column_dimensions['D'].width = 20
    ws2.column_dimensions['E'].width = 15
    ws2.column_dimensions['F'].width = 20
    ws2.column_dimensions['G'].width = 20
    ws2.column_dimensions['H'].width = 25

    r = 1
    ws2.cell(row=r, column=2, value="PAYMENT TRACKING").font = Font(bold=True, size=18, color="E8841A")
    ws2.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    r += 1
    ws2.cell(row=r, column=2, value=f"Reference: {data['reference']}  —  Client: {cl['company']}").font = Font(italic=True, color="666666")
    ws2.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    r += 2

    ws2.cell(row=r, column=2, value="Total Amount Due").font = bold_font
    ws2.cell(row=r, column=3, value=grand).number_format = '"$"#,##0'
    ws2.cell(row=r, column=3).font = bold_font
    r += 1
    ws2.cell(row=r, column=2, value="Total Paid (auto)").font = bold_font
    ws2.cell(row=r, column=3, value=f"=SUM(F{r+4}:F{r+20})").number_format = '"$"#,##0'
    ws2.cell(row=r, column=3).font = Font(bold=True, color="2E7D32")
    paid_cell_ref = f"C{r}"
    r += 1
    ws2.cell(row=r, column=2, value="Remaining Balance (auto)").font = bold_font
    ws2.cell(row=r, column=3, value=f"=C{r-2}-{paid_cell_ref}").number_format = '"$"#,##0'
    ws2.cell(row=r, column=3).font = Font(bold=True, color="CC0000")
    ws2.cell(row=r, column=3).fill = light_orange_fill
    r += 2

    headers = ["#", "Description", "Scheduled Date", "Scheduled Amount", "Paid Date", "Paid Amount", "Status", "Notes"]
    for col, h in enumerate(headers, start=2):
        c = ws2.cell(row=r, column=col, value=h)
        c.font = white_font; c.fill = orange_fill; c.alignment = center; c.border = border
    r += 1

    planning = [
        ("1", "First deposit", "", grand * 0.5, "", "", "", ""),
        ("2", "Final payment", "", grand * 0.5, "", "", "", ""),
    ]
    for i in range(3, 13):
        planning.append((str(i), "", "", "", "", "", "", ""))

    for p in planning:
        for col, val in enumerate(p, start=2):
            c = ws2.cell(row=r, column=col, value=val if val != "" else None)
            c.border = border
            if col in (5, 7):
                c.number_format = '"$"#,##0'
                c.alignment = right
            elif col in (2, 4, 6):
                c.alignment = center
        r += 1

    r += 1
    ws2.cell(row=r, column=2, value="How to use:").font = bold_font
    r += 1
    for line in [
        "• Fill in 'Scheduled Date' and 'Scheduled Amount' for each planned payment",
        "• When a payment is received, fill 'Paid Date' and 'Paid Amount'",
        "• Set Status: Scheduled / Partial / Paid / Overdue",
        "• 'Total Paid' and 'Remaining Balance' update automatically",
    ]:
        ws2.cell(row=r, column=2, value=line).font = Font(size=9, color="666666")
        ws2.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
        r += 1

    wb.save(out)
